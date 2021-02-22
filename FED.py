#!/usr/bin/python

import math as m
import numpy as np
import openpyxl as openpyxl

###############################################################################
# INPUTS
###############################################################################

L_tunel = 400 # [m]
N_sens = 200
dist_entre_sensores = 2.0 # [m]
vel_despl = 0.5 # [m/s]
t_previos = 150 # [s]
eval_time = 900 # [s]
position = 25

###############################################################################
# VARIABLE INITIATION
###############################################################################

ColsCO = [1,N_sens+1]
ColsCO2 = [N_sens+1,2*N_sens+1]
ColsO2 = [2*N_sens+1,3*N_sens+1]
ColsVIS = [3*N_sens+1,4*N_sens+1]
ColsTEMP = [4*N_sens+1,5*N_sens+1]

FED_CO = 0.0
HV_CO2 = 0.0
FED_O2 = 0.0
VIS = 0.0
TEMP = 0.0

###################################################################################
# FRACTIONAL EFFECTIVE DOSE FUNCTIONS
###################################################################################

# Hay que revisar si el argumento es % o decimales (21 vs 0.21)
def F_FED_CO (CO):

    return 2.764e-05 * CO**(1.036)


def F_FED_O2 (O2):

    return 1/(m.exp(8.13-0.54*(20.9-O2)))


def F_HV_CO2 (CO2):

    return m.exp(0.1903*CO2+2.0004)/7.1

def findIndex (Array,t):

    for i in range(len(Array)):

        if Array[i]<= t and Array[i+1]>= t:

            Index = i+1

            break

    return Index

###################################################################################
# FDS OUTPUT READ AND WRITE
###################################################################################

fds_output_file  = open('ASET_devc_5-7-19/ASET_devc_vp1_07.csv', "r")
fds_output_lines = fds_output_file.readlines()
fds_output_file.close()

fds_output = []
for line in fds_output_lines:
    aux1 = line.split(",")
    try:
        fds_output.append([float(k) for k in aux1])
    except:
        pass

fds_output = np.asarray(fds_output)

###################################################################################
# SPECIES MATRIX
###################################################################################

FDStime = fds_output[:,0]
Sens_dist = [i*dist_entre_sensores+1 for i in range(N_sens)]
C_CO = fds_output[:,ColsCO[0]:ColsCO[1]]
C_CO2 = fds_output[:,ColsCO2[0]:ColsCO2[1]]
C_O2 = fds_output[:,ColsO2[0]:ColsO2[1]]
DATOS_VIS = fds_output[:,ColsVIS[0]:ColsVIS[1]]
DATOS_TEMP = fds_output[:,ColsTEMP[0]:ColsTEMP[1]]

max_FDStime_index = findIndex(FDStime,eval_time)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'CÁLCULO'

sheet_dt = sheet.cell(1,1,'t [s]')
sheet_dt = sheet.cell(2,1,FDStime[0])
sheet_dt = sheet.cell(1,2,'dt [s]')
sheet_dt = sheet.cell(1,3,'Position [m]')
sheet_dt = sheet.cell(1,4,'FED_CO')
sheet_dt = sheet.cell(1,5,'FED_O2')
sheet_dt = sheet.cell(1,6,'HV_CO2')
sheet_dt = sheet.cell(1,7,'FED')
sheet_dt = sheet.cell(1,8,'VIS [m]')
sheet_dt = sheet.cell(1,9,'T [°C]')


for j in range(1,max_FDStime_index):

    sheet_t = sheet.cell(j+2,1,FDStime[j])
    dt = FDStime[j]-FDStime[j-1]
    sheet_dt = sheet.cell(j+2,2,dt)
    if FDStime[j] > t_previos:
        position+=dt*vel_despl
    if position > L_tunel:
        position = L_tunel
    sheet_dt = sheet.cell(j+2,3,position)
    col_no = findIndex(Sens_dist,position)
    FED_CO += F_FED_CO(1000000*C_CO[j,col_no]) * (dt/60.)
    sheet_dt = sheet.cell(j+2,4,FED_CO)
    FED_O2 += F_FED_O2(100*C_O2[j,col_no]) * (dt/60.)
    sheet_dt = sheet.cell(j+2,5,FED_O2)
    HV_CO2 = F_HV_CO2(100*C_CO2[j,col_no])
    sheet_dt = sheet.cell(j+2,6,HV_CO2)
    FED = FED_CO * HV_CO2 + FED_O2
    sheet_dt = sheet.cell(j+2,7,FED)
    VIS = DATOS_VIS[j,col_no]
    sheet_dt = sheet.cell(j+2,8,VIS)
    TEMP = DATOS_TEMP[j,col_no]
    sheet_dt = sheet.cell(j+2,9,TEMP)
    if position == L_tunel:
        break

wb.create_sheet('C_CO')
sheet = wb['C_CO']
for i in range(len(FDStime)):
    sheet2_t = sheet.cell(i+2,1,FDStime[i])
    for j in range(N_sens):
        sheet2_t = sheet.cell(i+2,j+2,C_CO[i,j])

wb.create_sheet('C_CO2')
sheet = wb['C_CO2']
for i in range(len(FDStime)):
    sheet2_t = sheet.cell(i+2,1,FDStime[i])
    for j in range(N_sens):
        sheet2_t = sheet.cell(i+2,j+2,C_CO2[i,j])

wb.create_sheet('C_O2')
sheet = wb['C_O2']
for i in range(len(FDStime)):
    sheet2_t = sheet.cell(i+2,1,FDStime[i])
    for j in range(N_sens):
        sheet2_t = sheet.cell(i+2,j+2,C_O2[i,j])

wb.create_sheet('TEMP')
sheet = wb['TEMP']
for i in range(len(FDStime)):
    sheet2_t = sheet.cell(i+2,1,FDStime[i])
    for j in range(N_sens):
        sheet2_t = sheet.cell(i+2,j+2,DATOS_TEMP[i,j])

wb.create_sheet('VIS')
sheet = wb['VIS']
for i in range(len(FDStime)):
    sheet2_t = sheet.cell(i+2,1,FDStime[i])
    for j in range(N_sens):
        sheet2_t = sheet.cell(i+2,j+2,DATOS_VIS[i,j])

wb.save('FED_vp1_07_150.xlsx')
